import openpyxl
import os
from persiantools.jdatetime import JalaliDate
import json
import datetime
import pandas as pd
#from mymailmerge import setWordData

def changeDatabase(exam_data,birth_date, grades,arzyab_name,first_name,last_name,
                    telephone_number,height,weight,hand,foot):
    path = os.getcwd() + '/MABC_2.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    ws = wb_obj.get_sheet_by_name(wb_obj.sheetnames[0])


    #--Birth-Date
    jyear = int(birth_date.split("/")[0])
    jmonth = int(birth_date.split("/")[1])
    jday = int(birth_date.split("/")[2])
    g = JalaliDate(jyear, jmonth, jday).to_gregorian()
    year = g.year
    month = g.month
    day = g.day

    ws["G2"] = year
    ws["F2"] = month
    ws["E2"] = day

    #--Exam-Date
    jyear = int(exam_data.split("/")[0])
    jmonth = int(exam_data.split("/")[1])
    jday = int(exam_data.split("/")[2])
    g = JalaliDate(jyear, jmonth, jday).to_gregorian()
    year = g.year
    month = g.month
    day = g.day

    ws["C2"] = year
    ws["B2"] = month
    ws["A2"] = day

    #Arzyab_Name
    ws["J2"] = arzyab_name

    #First_Name_Last_Name
    ws["K2"] = first_name
    ws["L2"] = last_name

    #telephone
    ws["P2"] = telephone_number

    #weight and height
    ws["Q2"] = height
    ws["R2"] = weight

    #hand foot
    ws["S2"] = hand
    ws["T2"] = foot

    grades_arr = json.loads(grades)

    ws["D5"] = grades_arr[0]
    ws["D6"] = grades_arr[1]
    ws["D7"] = grades_arr[2]
    ws["D8"] = grades_arr[3]
    ws["D9"] = grades_arr[4]
    ws["D10"] = grades_arr[5]
    ws["D11"] = grades_arr[6]
    ws["D12"] = grades_arr[7]
    ws["D13"] = grades_arr[8]
    ws["D14"] = grades_arr[9]
    ws["D15"] = grades_arr[10] if len(grades_arr) > 10 else 0
    
    wb_obj.save("MABC_2.xls")
    


def readGrades():
    # path = os.getcwd() + '/MABC_2.xlsx'
    # wb_obj = openpyxl.load_workbook(path,data_only=True)
    # ws = wb_obj.get_sheet_by_name(wb_obj.sheetnames[0])

    # return {
    #     "grade_1":              ws["G5"],
    #     "grade_2":              ws["G6"].value,
    #     "grade_3":              ws["G7"].value,
    #     "grade_4":              ws["G8"].value,
    #     "grade_5":              ws["G9"].value,
    #     "grade_6":              ws["G10"].value,
    #     "grade_7":              ws["G11"].value,
    #     "grade_8":              ws["G12"].value,
    #     "grade_9":              ws["G13"].value,
    #     "grade_10":             ws["G14"].value,
    #     "summation_of_total":   ws["N5"].value,
    #     "standard_grade_total": ws["O5"].value,
    #     "percentage_total":     ws["P5"].value,
    #     "age_group":            ws["M3"].value,
    #     "arzyab_name":          ws["J2"].value,
    #     "first_name":           ws["K2"].value,
    #     "last_name":            ws["L2"].value,
    #     "year":                 ws["N2"].value,
    #     "month":                ws["O2"].value,
    #     "telephone":            ws["P2"].value,
    #     "height":               ws["Q2"].value,
    #     "weight":               ws["R2"].value,
    #     "hand":                 ws["S2"].value,
    #     "foot":                 ws["T2"].value,
    #     "birth_date":           ws["H2"].value,
    #     "exam_date":            ws["D2"].value,
    #     "raw_grade_1":          ws["D5"].value,
    #     "raw_grade_2":          ws["D6"].value,
    #     "raw_grade_3":          ws["D7"].value,
    #     "raw_grade_4":          ws["D8"].value,
    #     "raw_grade_5":          ws["D9"].value,
    #     "raw_grade_6":          ws["D10"].value,
    #     "raw_grade_7":          ws["D11"].value,
    #     "raw_grade_8":          ws["D12"].value,
    #     "raw_grade_9":          ws["D13"].value,
    #     "raw_grade_10":         ws["D14"].value,
    #     "raw_grade_11":         ws["D15"].value,
    #     "MD1":                  ws["H18"].value,
    #     "MD2":                  ws["H19"].value,
    #     "MD3":                  ws["H20"].value,
    #     "A&C1":                 ws["H21"].value,
    #     "A&C2":                 ws["H22"].value,
    #     "Bal1":                 ws["H23"].value,
    #     "Bal2":                 ws["H24"].value,
    #     "Bal3":                 ws["H25"].value
    # }
    x = pd.read_excel("MABC_2.xlsx")
    print(x)

changeDatabase("1400/08/04","1397/08/03","[24.39,29.08,55.90,3,10,3,3.42,6.14,15,5]",
"هادی رسالتی","رزا","قاسم نعمتی","09361890427",130,40,"راست","راست")
(readGrades())